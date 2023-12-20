import openpyxl

def readData():
    list =[]
    path = "C://Users//shpaa//PycharmProjects//pythonProject//excel//credentials_python.xlsx"
    workbook =openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_row
    for r in range(2, rows+1):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple = (username, password)
        list.append(tuple)
    return list